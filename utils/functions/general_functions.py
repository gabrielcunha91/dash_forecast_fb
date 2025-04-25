import streamlit as st
import pandas as pd
import numpy as np
from utils.queries import *
from workalendar.america import Brazil
import openpyxl
import os
import re


def export_to_excel(df, sheet_name, excel_filename):
  if os.path.exists(excel_filename):
    wb = openpyxl.load_workbook(excel_filename)
  else:
    wb = openpyxl.Workbook()

  if sheet_name in wb.sheetnames:
    wb.remove(wb[sheet_name])
  ws = wb.create_sheet(title=sheet_name)
  
  # Escrever os cabeçalhos
  for col_idx, column_title in enumerate(df.columns, start=1):
    ws.cell(row=1, column=col_idx, value=column_title)
  
  # Escrever os dados
  for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
    for col_idx, value in enumerate(row, start=1):
      ws.cell(row=row_idx, column=col_idx, value=value)

  wb.save(excel_filename)


def format_brazilian(num):
  try:
    num = float(num)
    return f"{num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
  except (ValueError, TypeError):
    return num


def format_columns_brazilian(df, numeric_columns):
  for col in numeric_columns:
    if col in df.columns:
      df[col] = df[col].apply(format_brazilian)
  return df


def format_percentage(num):
  try:
    num = float(num)
    formatted_num = f"{num * 100:,.2f}"  # Multiplica por 100 e formata
    return f"{formatted_num.replace(',', 'X').replace('.', ',').replace('X', '.')}%"  # Formata como percentual
  except (ValueError, TypeError):
    return num  # Retorna o valor original em caso de erro
  

def format_columns_percentage(df, numeric_columns):
  for col in numeric_columns:
    if col in df.columns:
      df[col] = df[col].apply(format_percentage)
  return df


# Dataframe filtrado pela casa:
def df_filtrar_casa(df, id_casa):
    df_filtrado = df[df['Casa'] == id_casa]
    return df_filtrado


def df_filtrar_periodo_data(df, coluna_data, data_inicio, data_fim):

	data_inicio = pd.to_datetime(data_inicio)
	data_fim = pd.to_datetime(data_fim) + pd.DateOffset(days=1)

	df = df.copy()

	df[coluna_data] = pd.to_datetime(df[coluna_data])
	df_filtrado = df.loc[(df[coluna_data] >= data_inicio) & (df[coluna_data] < data_fim)]

	return df_filtrado

def is_valid_uuid(uuid):
    pattern = r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$'
    return bool(re.match(pattern, uuid))